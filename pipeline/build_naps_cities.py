"""
ANNUAL builder (NOT part of the weekly registry) for the city air-quality
dashboard — analogous to build_wait_times.py / build_geography.py.

Source: Environment and Climate Change Canada's National Air Pollution
Surveillance (NAPS) program. Open Government Licence – Canada. Two eras:
  * 2005–latest: the **Annual Summaries** product (pre-computed, validated
    per-station annual + 12 monthly means; CSV 2016+, .xlsx 2005–2015) — read the
    **1HR** block.
  * pre-2005: NAPS published only raw **hourly** files (no pre-computed means), so
    we aggregate them ourselves — mean of valid hourly values, requiring >=50% of a
    month's hours for a monthly mean and >=6 months across all 4 quarters for an
    annual mean. Reaches back to 1974 for the gases (NO2/SO2/O3/CO) and 1995 for
    PM2.5.
Stations are then averaged to each major city.

Output (committed CSVs, re-run each spring when validated NAPS lands):
  data/environment/naps_city_annual.csv    city, pollutant, unit, year, value, n_stations
  data/environment/naps_city_monthly.csv   city, pollutant, unit, year, month, date, value

Run:  python -m pipeline.build_naps_cities
"""

import io
import csv
import re
import calendar
import unicodedata
import requests
import openpyxl
import pandas as pd
from datetime import date
from pipeline.config import DATA_DIR
from pipeline.metadata import save_metadata
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

_UA = {"User-Agent": "Mozilla/5.0 (DataCan pipeline; +https://canadaobservatory.github.io)"}
_FILE = "https://data-donnees.az.ec.gc.ca/api/file?path="
_NAPS = ("/air/monitor/national-air-pollution-surveillance-naps-program/Data-Donnees")

START_YEAR = 2005   # AnnualSummaries (per-station means) begin here
# Continuous pollutants to chart: token in filename -> (display, units)
POLLUTANTS = {
    "PM25": ("PM2.5", "µg/m³"),
    "O3":   ("O3", "ppb"),
    "NO2":  ("NO2", "ppb"),
    "SO2":  ("SO2", "ppb"),
    "CO":   ("CO", "ppm"),
}

# Curated major cities (CMA-level). Each: display -> (province set, match substrings).
# Province-gated so e.g. Windsor (ON) doesn't catch "Grand Falls - Windsor" (NL).
# A city with no NAPS data simply won't appear in the output.
CANONICAL = [
    ("Toronto",            {"ON"}, ["toronto"]),
    ("Montréal",           {"QC"}, ["montréal", "montreal"]),
    ("Vancouver",          {"BC"}, ["vancouver"]),          # incl. "Metro Vancouver - *"
    ("Calgary",            {"AB"}, ["calgary"]),
    ("Edmonton",           {"AB"}, ["edmonton"]),
    ("Ottawa–Gatineau",    {"ON", "QC"}, ["ottawa", "gatineau"]),
    ("Winnipeg",           {"MB"}, ["winnipeg"]),
    ("Québec City",        {"QC"}, ["québec", "quebec"]),
    ("Hamilton",           {"ON"}, ["hamilton"]),
    ("Kitchener–Waterloo", {"ON"}, ["kitchener", "waterloo"]),
    ("London",             {"ON"}, ["london"]),
    ("Windsor",            {"ON"}, ["windsor"]),
    ("St. Catharines–Niagara", {"ON"}, ["catharines", "niagara"]),
    ("Sudbury",            {"ON"}, ["sudbury"]),
    ("Thunder Bay",        {"ON"}, ["thunder bay"]),
    ("Halifax",            {"NS"}, ["halifax", "dartmouth"]),
    ("Saint John",         {"NB"}, ["saint john"]),
    ("Fredericton",        {"NB"}, ["fredericton"]),
    ("Moncton",            {"NB"}, ["moncton"]),
    ("Charlottetown",      {"PE"}, ["charlottetown"]),
    ("St. John's",         {"NL"}, ["st. john's", "st john's", "mount pearl"]),
    ("Victoria",           {"BC"}, ["victoria"]),
    ("Kelowna",            {"BC"}, ["kelowna"]),
    ("Abbotsford",         {"BC"}, ["abbotsford"]),
    ("Regina",             {"SK"}, ["regina"]),
    ("Saskatoon",          {"SK"}, ["saskatoon"]),
    ("Sherbrooke",         {"QC"}, ["sherbrooke"]),
    ("Saguenay",           {"QC"}, ["saguenay", "jonquière", "chicoutimi"]),
    ("Trois-Rivières",     {"QC"}, ["trois-rivières", "trois-rivieres"]),
    ("Whitehorse",         {"YT"}, ["whitehorse"]),
    ("Yellowknife",        {"NT"}, ["yellowknife"]),
]


_PT_ALIAS = {"PQ": "QC", "NF": "NL"}   # historical province codes -> modern


def _norm(s):
    """Lowercase, strip accents + punctuation so old/new spellings match
    ('St Johns' vs "St. John's", 'Montreal' vs 'Montréal')."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = s.replace("'", "")                 # drop apostrophes: john's -> johns
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", s)).strip()


def _canon(city, pt):
    pt = _PT_ALIAS.get(str(pt or "").strip().upper(), str(pt or "").strip().upper())
    cl = _norm(city)
    for display, provs, pats in CANONICAL:
        if pt in provs and any(_norm(p) in cl for p in pats):
            return display
    return None


def _fnum(x):
    try:
        v = float(x)
        return v if v > -999 else None
    except (TypeError, ValueError):
        return None


def _parse_1hr_block(text):
    """Return (header_index_map, list-of-raw-rows) for the 1HR block, or (None, [])."""
    lines = text.splitlines()
    starts = [i for i, l in enumerate(lines) if re.search(r"\*{3,}\s*1HR\s*\*{3,}", l)]
    if not starts:
        return None, []
    h = starts[0] + 1
    ends = [i for i, l in enumerate(lines)
            if i > h and re.search(r"\*{3,}\s*(8HR|24HR|DMAX1HR|DAILYMEAN)\s*\*{3,}", l)]
    end = ends[0] if ends else len(lines)
    hdr = next(csv.reader([lines[h]]))
    ci = {n.strip(): i for i, n in enumerate(hdr)}
    if "City" not in ci or "Mean" not in ci or "P/T" not in ci:
        return None, []
    rows = [r for r in csv.reader(lines[h + 1:end]) if r and r[0].strip()]
    return ci, rows


def _parse_xlsx_1hr(content):
    """Pre-2016 years ship as .xlsx with one sheet per averaging basis. Read the
    '1hr' sheet -> (header_index_map, list-of-row-lists)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.lower() == "1hr"), None)
    if sheet is None:
        return None, []
    rows = list(wb[sheet].iter_rows(values_only=True))
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if "City" in cells and "Mean" in cells and "P/T" in cells:
            ci = {n: j for j, n in enumerate(cells)}
            data = [list(rr) for rr in rows[i + 1:] if rr and rr[ci["City"]] is not None]
            return ci, data
    return None, []


def _rows_for(tok, year):
    """Fetch one pollutant-year's 1-hour station table, trying the modern CSV
    (2016+) then the older .xlsx. Returns (header_index_map, rows) or (None, [])."""
    base = (f"{_FILE}{_NAPS}/{year}/ContinuousData-DonneesContinu/"
            f"AnnualSummaries-SommairesAnnuels/{year}_Annual{tok}")
    r = requests.get(base + "_EN.csv", headers=_UA, timeout=60)
    if r.status_code == 200:
        return _parse_1hr_block(r.content.decode("utf-8-sig", errors="replace"))
    r = requests.get(base + ".xlsx", headers=_UA, timeout=90)
    if r.status_code == 200:
        return _parse_xlsx_1hr(r.content)
    return None, []


# --- pre-2005 back-build from raw hourly files (no pre-computed means exist) ---
PRE = 2005                              # hourly for years < PRE; Annual Summaries for >= PRE
START_HOURLY = {"PM25": 1995, "O3": 1974, "NO2": 1974, "SO2": 1974, "CO": 1974}


def _hourly_url(tok, year):
    return (f"{_FILE}{_NAPS}/{year}/ContinuousData-DonneesContinu/"
            f"HourlyData-DonneesHoraires/{tok}_{year}.csv")


def _aggregate_hourly(text, year):
    """Parse one pollutant-year HOURLY file (4-line preamble, then header
    'Pollutant,NAPSID,City,P/T,Lat,Lon,Date,H01..H24'; -999 = missing) into
    per-station monthly means (>=50% of the month's hours valid) and an annual mean
    (>=6 months spanning all 4 quarters), each mapped to a canonical city.
    Returns list of (city, month_or_0, value); month 0 = the annual mean."""
    rdr = list(csv.reader(text.splitlines()))
    hidx = next((i for i, r in enumerate(rdr[:40])
                 if "NAPSID" in r and "Date" in r and "H01" in r), None)
    if hidx is None:
        return []
    ci = {n.strip(): j for j, n in enumerate(rdr[hidx])}
    hcols = [ci[f"H{h:02d}"] for h in range(1, 25) if f"H{h:02d}" in ci]
    ic, ip, inn, idt = ci.get("City"), ci.get("P/T"), ci.get("NAPSID"), ci.get("Date")
    if None in (ic, ip, inn, idt) or not hcols:
        return []
    from collections import defaultdict
    sm = defaultdict(list)      # (naps_id, month) -> [valid hourly values]
    meta = {}                   # naps_id -> (city, pt)
    for r in rdr[hidx + 1:]:
        if len(r) <= idt or not r[inn].strip():
            continue
        d = r[idt].strip()
        if len(d) < 6 or not d[:6].isdigit():
            continue
        mo = int(d[4:6])
        if not 1 <= mo <= 12:
            continue
        nid = r[inn].strip()
        meta[nid] = (r[ic], r[ip])
        bucket = sm[(nid, mo)]
        for hc in hcols:
            if hc < len(r):
                v = r[hc].strip()
                if v and v not in ("-999", "-999.0"):
                    try:
                        fv = float(v)
                        if fv > -900:
                            bucket.append(fv)
                    except ValueError:
                        pass
    out = []
    for nid in {n for (n, _m) in sm}:
        city = _canon(*meta.get(nid, ("", "")))
        if city is None:
            continue
        mmeans = {}
        for mo in range(1, 13):
            vals = sm.get((nid, mo), [])
            if vals and len(vals) >= 0.5 * calendar.monthrange(year, mo)[1] * 24:
                mmeans[mo] = sum(vals) / len(vals)
        for mo, mv in mmeans.items():
            out.append((city, mo, mv))
        if len(mmeans) >= 6 and len({(mo - 1) // 3 for mo in mmeans}) == 4:
            out.append((city, 0, sum(mmeans.values()) / len(mmeans)))
    return out


def build_naps_cities():
    end_year = date.today().year
    # accumulators: {(city, poll): {year: [means]}} and monthly {(city,poll,year,month):[vals]}
    ann = {}      # (city, poll) -> {year: [station means]}
    mon = {}      # (city, poll, year, monthnum) -> [station monthly means]
    units = {}

    for tok, (poll, unit) in POLLUTANTS.items():
        units[poll] = unit
        got_years = 0
        for year in range(START_YEAR, end_year + 1):
            try:
                ci, rows = _rows_for(tok, year)
            except Exception as e:
                logger.warning(f"  {poll} {year}: {e}")
                continue
            if ci is None:
                continue
            mean_i = ci["Mean"]
            month_idx = list(range(mean_i - 12, mean_i))   # 12 cols before Mean = Jan..Dec
            for r in rows:
                try:
                    pt = str(r[ci["P/T"]] or "").strip()
                    city = _canon(r[ci["City"]], pt)
                except (IndexError, KeyError):
                    continue
                if city is None:
                    continue
                m = _fnum(r[mean_i]) if mean_i < len(r) else None
                if m is not None:
                    ann.setdefault((city, poll), {}).setdefault(year, []).append(m)
                for mi, col in enumerate(month_idx, start=1):
                    v = _fnum(r[col]) if col < len(r) else None
                    if v is not None:
                        mon.setdefault((city, poll, year, mi), []).append(v)
            got_years += 1
        logger.info(f"  {poll}: read {got_years} years (2005+, Annual Summaries)")

        # pre-2005: aggregate from raw hourly files (no pre-computed means exist)
        hstart = START_HOURLY.get(tok)
        h_years = 0
        if hstart:
            for year in range(hstart, PRE):
                try:
                    rr = requests.get(_hourly_url(tok, year), headers=_UA, timeout=240)
                    if rr.status_code != 200:
                        continue
                    recs = _aggregate_hourly(rr.content.decode("latin-1", errors="replace"), year)
                except Exception as e:
                    logger.warning(f"  {poll} {year} (hourly): {e}")
                    continue
                for city, mo, val in recs:
                    if mo == 0:
                        ann.setdefault((city, poll), {}).setdefault(year, []).append(val)
                    else:
                        mon.setdefault((city, poll, year, mo), []).append(val)
                if recs:
                    h_years += 1
            logger.info(f"  {poll}: + {h_years} pre-2005 years (raw hourly)")

    # ---- annual frame ----
    a_rows = []
    for (city, poll), yd in ann.items():
        for year, vals in yd.items():
            a_rows.append({"city": city, "pollutant": poll, "unit": units[poll],
                           "year": year, "value": round(sum(vals) / len(vals), 2),
                           "n_stations": len(vals)})
    adf = pd.DataFrame(a_rows).sort_values(["pollutant", "city", "year"]).reset_index(drop=True)

    # ---- monthly frame ----
    m_rows = []
    for (city, poll, year, mi), vals in mon.items():
        m_rows.append({"city": city, "pollutant": poll, "unit": units[poll],
                       "year": year, "month": mi,
                       "date": f"{year}-{mi:02d}-01",
                       "value": round(sum(vals) / len(vals), 2)})
    mdf = pd.DataFrame(m_rows).sort_values(["pollutant", "city", "year", "month"]).reset_index(drop=True)

    if adf.empty:
        logger.error("NAPS city build produced no data")
        return None

    (DATA_DIR / "environment").mkdir(parents=True, exist_ok=True)
    a_out = DATA_DIR / "environment" / "naps_city_annual.csv"
    m_out = DATA_DIR / "environment" / "naps_city_monthly.csv"
    adf.to_csv(a_out, index=False)
    mdf.to_csv(m_out, index=False)
    for out, df in [(a_out, adf), (m_out, mdf)]:
        save_metadata(out, df=df, latest_observation_date=str(int(adf["year"].max())),
            source="Environment and Climate Change Canada",
            source_table="ECCC National Air Pollution Surveillance (NAPS) — Annual Summaries (2005–) + raw hourly back-build (pre-2005), 1-hour basis",
            frequency="annual", unit="µg/m³ (PM) / ppb (gases) / ppm (CO)",
            transformations=["1-hour-basis station means aggregated to major cities",
                             "annual + 12 monthly means; 2005– from Annual Summaries, pre-2005 from raw hourly (gases to 1974, PM2.5 to 1995)"])
    logger.info(f"  cities: {adf.city.nunique()} | annual rows {len(adf)} | monthly rows {len(mdf)}")
    logger.info(f"  saved -> {a_out.name}, {m_out.name}")
    return adf, mdf


if __name__ == "__main__":
    build_naps_cities()
