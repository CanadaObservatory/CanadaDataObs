"""
Geography-section registry fetchers (run weekly by run_pipeline, source="custom"):

  fetch_wildfire  — annual area burned, national + by province/territory
                    (NRCan Canadian National Fire Database, point-data summary)
  fetch_sea_ice   — Arctic sea-ice extent by month, 1979–present
                    (NSIDC Sea Ice Index, Version 4, G02135)

The section's STATIC map assets (province/ecozone/permafrost boundaries, density,
freshwater, land cover) are built once by pipeline/build_geography.py, not here.
"""

import io
import requests
import pandas as pd
from pipeline.config import DATA_DIR
from pipeline.metadata import save_metadata
import logging

logger = logging.getLogger(__name__)

# --- Wildfire (NRCan NFDB) ---------------------------------------------------
NFDB_XLSX = ("https://cwfis.cfs.nrcan.gc.ca/downloads/nfdb/fire_pnt/"
             "current_version/NFDB_point_stats.xlsx")
# NFDB agency codes -> province/territory name. "PC" (Parks Canada federal lands)
# is excluded from the province map but is included in the national total; "NU"
# has no NFDB point data (no TOTAL_HA column → absent; wildfire there is negligible).
AGENCY_PROV = {
    "AB": "Alberta", "BC": "British Columbia", "MB": "Manitoba",
    "NB": "New Brunswick", "NL": "Newfoundland and Labrador", "NS": "Nova Scotia",
    "NT": "Northwest Territories", "ON": "Ontario", "PEI": "Prince Edward Island",
    "QC": "Quebec", "SK": "Saskatchewan", "YT": "Yukon",
}


def _year_rows(grid, year_col=0):
    """Yield rows of an openpyxl values grid whose year cell is a real 4-digit year
    (skips the title/blank rows and the trailing '10 yr average' summary rows)."""
    for r in grid:
        y = r[year_col] if year_col < len(r) else None
        if isinstance(y, (int, float)) and 1900 <= y <= 2100:
            yield int(y), r


def fetch_wildfire():
    """Annual area burned (hectares): national series from the summary sheet, plus
    a by-province/territory series parsed from the wide 'by agency' sheet."""
    logger.info("Fetching NRCan NFDB wildfire area burned...")
    try:
        import openpyxl
        r = requests.get(NFDB_XLSX, timeout=180)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"  NFDB fetch failed: {e}")
        return None

    rows = []   # (geography, year, area_burned_ha)

    # national: NFDB_Summary_Stats, header row begins with "YEAR"; col TOTAL_HA
    ws = list(wb["NFDB_Summary_Stats"].iter_rows(values_only=True))
    h = next(i for i, x in enumerate(ws) if x and x[0] == "YEAR")
    cols = list(ws[h])
    ti = cols.index("TOTAL_HA")
    for y, r in _year_rows(ws[h + 1:]):
        rows.append(("Canada", y, r[ti] if ti < len(r) else None))

    # by province: NFDB_Summary_Stats_By_Agency. Agency codes sit two rows above
    # the "YEAR" sub-header; forward-fill the code across its block and take that
    # block's exact TOTAL_HA column (robust to the irregular Nunavut block).
    wsa = list(wb["NFDB_Summary_Stats_By_Agency"].iter_rows(values_only=True))
    ha = next(i for i, x in enumerate(wsa) if x and x[0] == "YEAR")
    code_row, sub_row = wsa[ha - 2], wsa[ha]
    colprov, cur = {}, None
    for j, v in enumerate(code_row):
        if v:
            cur = str(v).strip()
        if cur in AGENCY_PROV and j < len(sub_row) and sub_row[j] == "TOTAL_HA":
            colprov[AGENCY_PROV[cur]] = j
    for y, r in _year_rows(wsa[ha + 1:]):
        for prov, j in colprov.items():
            rows.append((prov, y, r[j] if j < len(r) else None))

    df = pd.DataFrame(rows, columns=["geography", "year", "area_burned_ha"])
    df["area_burned_ha"] = pd.to_numeric(df["area_burned_ha"], errors="coerce")
    df = (df.dropna(subset=["area_burned_ha"])
            .sort_values(["geography", "year"]).reset_index(drop=True))
    if df.empty:
        return None

    out_path = DATA_DIR / "geography" / "nfdb_wildfire.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_path, index=False)
    y0, y1 = int(df.year.min()), int(df.year.max())
    save_metadata(out_path, df=df, latest_observation_date=str(y1),
        source="Natural Resources Canada / Canadian Forest Service",
        source_table="Canadian National Fire Database (NFDB), point-data summary",
        frequency="annual", unit="hectares burned",
        transformations=[f"national TOTAL_HA + by-agency TOTAL_HA; {y0}–{y1}; "
                         "Parks Canada folded into national only; Nunavut has no point data"])
    logger.info(f"  saved {len(df)} rows ({df.geography.nunique()} geographies, {y0}–{y1}) "
                f"-> {out_path.name}")
    return df


# --- Sea ice (NSIDC Sea Ice Index, G02135) ----------------------------------
NSIDC_MONTHLY = ("https://noaadata.apps.nsidc.org/NOAA/G02135/north/monthly/data/"
                 "N_{:02d}_extent_v4.0.csv")


def fetch_sea_ice():
    """Arctic (Northern Hemisphere) monthly sea-ice extent, 1979–present. The
    page charts the September minimum, the clearest long-run climate signal."""
    logger.info("Fetching NSIDC Arctic sea-ice extent (G02135)...")
    frames = []
    for mo in range(1, 13):
        try:
            r = requests.get(NSIDC_MONTHLY.format(mo), timeout=60)
            r.raise_for_status()
        except Exception as e:
            logger.warning(f"  sea-ice month {mo:02d} failed: {e}")
            continue
        d = pd.read_csv(io.StringIO(r.text))
        d.columns = [c.strip() for c in d.columns]      # header has stray spaces
        frames.append(d.rename(columns={"mo": "month", "extent": "extent_mkm2",
                                         "area": "area_mkm2"})
                      [["year", "month", "extent_mkm2", "area_mkm2"]])
    if not frames:
        logger.error("  sea ice: no monthly files fetched")
        return None

    df = pd.concat(frames, ignore_index=True)
    for c in ("year", "month", "extent_mkm2", "area_mkm2"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df[df["extent_mkm2"] > 0]                       # drop -9999 missing flag
    df["year"] = df["year"].astype(int)
    df["month"] = df["month"].astype(int)
    df = df.sort_values(["month", "year"]).reset_index(drop=True)
    if df.empty:
        return None

    out_path = DATA_DIR / "geography" / "nsidc_sea_ice.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_path, index=False)
    save_metadata(out_path, df=df, latest_observation_date=str(int(df.year.max())),
        source="NSIDC Sea Ice Index, Version 4 (G02135), NSIDC / NOAA",
        source_table="G02135 monthly Northern Hemisphere sea-ice extent",
        frequency="monthly", unit="million km² (sea-ice extent)",
        transformations=["12 monthly NH extent files concatenated; dropped -9999 (missing)"])
    logger.info(f"  saved {len(df)} rows ({int(df.year.min())}–{int(df.year.max())}) "
                f"-> {out_path.name}")
    return df


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    fetch_wildfire()
    fetch_sea_ice()
